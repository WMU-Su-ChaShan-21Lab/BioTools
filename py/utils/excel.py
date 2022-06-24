# -*- encoding: utf-8 -*-
"""
@File Name      :   excel.py    
@Create Time    :   2022/1/11 9:48
@Description    :   
@Version        :   
@License        :   MIT
@Author         :   diklios
@Contact Email  :   diklios5768@gmail.com
@Github         :   https://github.com/diklios5768
@Blog           :   
@Motto          :   All our science, measured against reality, is primitive and childlike - and yet it is the most precious thing we have.
"""
__auth__ = 'diklios'

from openpyxl import Workbook, load_workbook

from utils.file import remove_file


def read_xlsx(file_path, sheet_name, column_index):
    wb = load_workbook(filename=file_path, data_only=True)
    ws = wb[sheet_name]
    table_data = []
    rows = ws.rows
    for row in rows:
        row_data = []
        col = row[column_index]
        if col.value:
            row_data.append(col.value)
        if row_data:
            table_data.append(row_data)
    return table_data


def generate_xlsx_file(filename, table_sheet):
    wb = Workbook()
    ws = wb.create_sheet(table_sheet['name'])
    table_data = table_sheet['data']
    rows_length = len(table_data)
    for i in range(rows_length):
        col_length = len(table_data[i])
        for j in range(col_length):
            ws.cell(row=i + 1, column=j + 1, value=table_data[i][j])
    wb.remove(wb['Sheet'])
    remove_file(filename)
    wb.save(filename)

