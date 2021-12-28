import os
import re
import requests
from openpyxl import Workbook, load_workbook
from requests.adapters import HTTPAdapter
from urllib.parse import quote
from urllib3.util.retry import Retry


session = requests.Session()
# 设置失败后进行重访问，最大5次
session.mount('https://', HTTPAdapter(max_retries=Retry(total=5,
              allowed_methods=frozenset(['GET', 'POST']))))

base_url = 'https://eutils.ncbi.nlm.nih.gov/entrez/eutils'
search_url = base_url + '/esearch.fcgi?db=snp&term='
summary_url = base_url + '/esummary.fcgi?db=snp&id='
fetch_url = base_url + '/efetch.fcgi?db=snp&id='
api_key = ''


def get_rs_id(gene_name):
    search_term = str(
        gene_name) + '[Gene Name] AND (likely pathogenic[Clinical_Significance] OR pathogenic[Clinical_Significance])'
    search_term = quote(search_term, encoding='utf-8')
    final_url = search_url + search_term + '&retmax=1000000&retmode=json'
    if api_key:
        final_url += '&api_key=' + api_key
    r = session.get(final_url)
    r.raise_for_status()
    search_results = r.json()
    return search_results['esearchresult']['idlist']


def get_rs_id_details(rs_id):
    final_url = summary_url + rs_id + '&retmode=json'
    if api_key:
        final_url += '&api_key=' + api_key
    r = session.get(final_url)
    r.raise_for_status()
    search_results = r.json()
    uid = search_results['result']['uids'][0]
    return search_results['result'][uid]


def group(list_to_group, step=100):
    return [list_to_group[i:i + step]
            for i in range(0, len(list_to_group), step)]


def get_rs_id_details_all(rs_id_list):
    step = 100
    rs_id_list_group = group(rs_id_list, step)
    rows = []
    for rs_id_list_10 in rs_id_list_group:
        final_url = summary_url + ','.join(rs_id_list_10) + \
            '&retmode=json'
        if api_key:
            final_url += '&api_key=' + api_key
        r = session.get(final_url)
        r.raise_for_status()
        search_results = r.json()
        # print(search_results)
        uids = search_results['result']['uids']
        rows.extend([search_results['result'][uid] for uid in uids])
    return rows


def handle_rs_id_details(rs_id_details):
    row = {}
    chr_pos = rs_id_details['chrpos']
    chromosome, pos = chr_pos.split(':')
    row['#chr'] = chromosome
    row['pos'] = pos
    row['variation'] = ';'.join([':'.join(item.split(':')[-2:])
                                 for item in rs_id_details['spdi'].split(',')])
    row['variant_type'] = rs_id_details['snp_class']
    row['snp_id'] = str(rs_id_details['snp_id'])
    row['clinical_significance'] = rs_id_details['clinical_significance']
    row['validation_status'] = rs_id_details['validated']
    row['function_class'] = ';'.join(rs_id_details['fxn_class'].split(','))
    row['gene'] = rs_id_details['genes'][0]['name']
    global_mafs = rs_id_details.get('global_mafs', '')
    if global_mafs:
        frequency_list = [
            re.sub(
                r'[=/]',
                ':',
                item['freq']) +
            ':' +
            item['study'] for item in global_mafs]
        row['frequency'] = '|'.join(frequency_list)
    else:
        row['frequency'] = ''
    return row


def handle_rs_id_details_to_table_data(name='snp', snp_list=None):
    if snp_list is None:
        snp_list = [{}]
    table_sheets = {'name': name}
    table_data = [['#chr',
                   'pos',
                   'variation',
                   'variant_type',
                   'snp_id',
                   'clinical_significance',
                   'validation_status',
                   'function_class',
                   'gene',
                   'frequency']]
    table_data.extend([list(snp.values()) for snp in snp_list])
    table_sheets['data'] = table_data
    return table_sheets


def fetch_rs_id_details(rs_id):
    final_url = fetch_url + rs_id + '&rettype=json&retmode=text'
    if api_key:
        final_url += '&api_key=' + api_key
    r = session.get(final_url)
    r.raise_for_status()
    search_results = r.json()
    # print(search_results)
    return search_results


def fetch_rs_id_details_all(rs_id_list):
    final_url = fetch_url + ','.join(rs_id_list) + '&rettype=json&retmode=text'
    if api_key:
        final_url += '&api_key=' + api_key
    r = session.get(final_url)
    r.raise_for_status()
    search_results = r.json()
    # print(search_results)
    return search_results


def read_xlsx(file_path, sheet_name, column_index):
    wb = load_workbook(filename=file_path, data_only=True)
    ws = wb[sheet_name]
    table_data = []
    rows = ws.rows
    for row in rows:
        row_data = []
        col = row[column_index]
        if col.value:
            row_data.append(col.value)
        if row_data:
            table_data.append(row_data)
    return table_data


def remove_file(file_path):
    """
    有文件就删除文件，没有文件就什么都不做
    """
    if os.path.isfile(file_path):
        os.remove(file_path)
        return True
    else:
        return False


def generate_xlsx_file(filename, table_sheet):
    wb = Workbook()
    ws = wb.create_sheet(table_sheet['name'])
    table_data = table_sheet['data']
    rows_length = len(table_data)
    for i in range(rows_length):
        col_length = len(table_data[i])
        for j in range(col_length):
            ws.cell(row=i + 1, column=j + 1, value=table_data[i][j])
    wb.remove(wb['Sheet'])
    if remove_file(filename):
        wb.save(filename)
        return True
    else:
        return False


if __name__ == '__main__':
    """
    自己写需要修改的参数：
    1.读取表格的路径、表格的sheet名称、表格的列索引
    2.生成表的路径、表的名称
    3.查询规则，详见官网：https://www.ncbi.nlm.nih.gov/snp/docs/eutils_help/，https://www.ncbi.nlm.nih.gov/snp/advanced
    4.个人的api_key，也可以不用填写，但是这样可以防止因为一秒钟访问次数过多被封禁
    """
    rs_id_details_list_all = []
    gene_names_table_data = read_xlsx('高度近视panel.xlsx', 'total1110Genes', 23)
    gene_names = [gene_name[0] for gene_name in gene_names_table_data]
    for gene in gene_names:
        rs_id_list_data = get_rs_id(gene)
        rs_id_details_list = get_rs_id_details_all(rs_id_list_data)
        rs_id_details_list_all.extend(rs_id_details_list)
    table_sheets_data = handle_rs_id_details_to_table_data(
        'snp', [
            handle_rs_id_details(rs_id_details) for rs_id_details in rs_id_details_list_all])
    generate_xlsx_file('rs_id_details.xlsx', table_sheets_data)
